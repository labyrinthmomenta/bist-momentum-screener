"""
engine.py — Momentum & FIP hesaplama motoru
Wesley R. Gray — Quantitative Momentum metodolojisi
Excel'den bağımsız, tamamen dinamik hesaplama.
"""
import math
import datetime
import openpyxl
from pathlib import Path


def _sf(v):
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except:
        return None


def _month_window(year, month):
    """Bir ayın ilk ve son takvim gününü döndür."""
    first = datetime.date(year, month, 1)
    if month == 12:
        last = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        last = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
    return first, last


def _prev_month(year, month):
    if month == 1:
        return year - 1, 12
    return year, month - 1


def load_raw_data(excel_path: Path) -> dict:
    """
    Excel'den ham günlük % getiri verisini yükler.
    Döner: {
        'tickers':        {ticker: {'type', 'industry', 'name'}},
        'daily':          {ticker: {date: return_decimal}},
        'dates_sorted':   [date, ...],
        'last_date':      date,
    }
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # ── Tarih indeksi ────────────────────────────────────────────────────────
    ws_raw  = wb['BIST D Return Data']
    date_row = list(ws_raw.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    idx_to_date = {
        i: v.date()
        for i, v in enumerate(date_row)
        if isinstance(v, datetime.datetime)
    }

    # ── Trade type ve industry ───────────────────────────────────────────────
    meta = {}   # ticker -> {type, industry, name}
    ws_ms = wb['MOMENTUM SCREENER']
    for row in ws_ms.iter_rows(min_row=5, max_row=650, values_only=True):
        if not row[4]: continue
        ticker = str(row[4]).strip()
        meta[ticker] = {
            'type':     str(row[1]).strip() if row[1] else 'Stock',
            'industry': str(row[2]).strip() if row[2] else '',
            'name':     str(row[5]).strip() if row[5] else ticker,
        }

    # ── Günlük getiri verisi ─────────────────────────────────────────────────
    daily = {}
    for row in ws_raw.iter_rows(min_row=6, max_row=650, values_only=True):
        ticker = row[3]
        if not ticker: continue
        ticker = str(ticker).strip()
        d = {}
        for idx, date in idx_to_date.items():
            val = _sf(row[idx]) if idx < len(row) else None
            if val is not None:
                d[date] = val / 100.0   # yüzde → ondalık
        daily[ticker] = d

    wb.close()

    all_dates = sorted(idx_to_date.values())
    filled_dates = [
        date for date in all_dates
        if any(date in daily[t] for t in daily)
    ]
    last_date = max(filled_dates) if filled_dates else None

    return {
        'tickers':      meta,
        'daily':        daily,
        'dates_sorted': all_dates,
        'last_date':    last_date,
    }


# ── Momentum ─────────────────────────────────────────────────────────────────

def calc_momentum_12_1(daily: dict, as_of_date: datetime.date):
    """
    Gray 12-1M: as_of_date'deki cari ayı hariç tut,
    önceki 12 ayın kümülatif getirisini hesapla.
    Örn: Şubat 2026 → Şubat 2025 başı – Ocak 2026 sonu
    """
    y, m = as_of_date.year, as_of_date.month
    # Cari ayı atla → bir önceki ay
    py, pm = _prev_month(y, m)
    _, window_end = _month_window(py, pm)        # Ocak 2026'nın son günü
    # 12 ay geriye git
    sm = pm - 11
    sy = py
    while sm <= 0:
        sm += 12
        sy -= 1
    window_start, _ = _month_window(sy, sm)     # Şubat 2025'in ilk günü

    product = 1.0
    count   = 0
    for date, ret in daily.items():
        if window_start <= date <= window_end:
            product *= (1.0 + ret)
            count   += 1
    return (product - 1.0) if count >= 10 else None


def calc_monthly_momentum(daily: dict, year: int, month: int):
    """Belirli bir ayın içi momentum değeri (ay bitmeden günceli gösterir)."""
    first, last = _month_window(year, month)
    product = 1.0
    count   = 0
    for date, ret in daily.items():
        if first <= date <= last:
            product *= (1.0 + ret)
            count   += 1
    return (product - 1.0) if count >= 1 else None


# ── FIP ──────────────────────────────────────────────────────────────────────

def calc_fip_annual(daily: dict, as_of_date: datetime.date):
    """
    Yıllık FIP: sign(12-1M) × (neg% − pos%)
    12-1M ile aynı pencere kullanılır.
    """
    y, m = as_of_date.year, as_of_date.month
    py, pm = _prev_month(y, m)
    _, window_end = _month_window(py, pm)
    sm = pm - 11
    sy = py
    while sm <= 0:
        sm += 12
        sy -= 1
    window_start, _ = _month_window(sy, sm)

    rets = [r for d, r in daily.items() if window_start <= d <= window_end]
    if len(rets) < 10: return None

    mom = calc_momentum_12_1(daily, as_of_date)
    if mom is None: return None

    total     = len(rets)
    neg_pct   = sum(1 for r in rets if r < 0) / total
    pos_pct   = sum(1 for r in rets if r > 0) / total
    sign      = 1 if mom >= 0 else -1
    return round(sign * (neg_pct - pos_pct), 6)


def calc_fip_monthly(daily: dict, year: int, month: int):
    """
    Aylık FIP: sign(aylık mom) × (neg% − pos%)
    """
    first, last = _month_window(year, month)
    rets = [r for d, r in daily.items() if first <= d <= last]
    if not rets: return None

    mom = calc_monthly_momentum(daily, year, month)
    if mom is None: return None

    total   = len(rets)
    neg_pct = sum(1 for r in rets if r < 0) / total
    pos_pct = sum(1 for r in rets if r > 0) / total
    sign    = 1 if mom >= 0 else -1
    return round(sign * (neg_pct - pos_pct), 6)


# ── Ana hesaplama ─────────────────────────────────────────────────────────────

def compute_all(raw: dict) -> list:
    """
    Tüm hisseler için momentum + FIP hesaplar.
    Döner: sorted list of stock dicts (12-1M mom'a göre azalan)
    """
    last_date = raw['last_date']
    y, m = last_date.year, last_date.month

    # Son 13 aylık veriyi hesapla
    months_back = []
    cy, cm = y, m
    for _ in range(13):
        months_back.append((cy, cm))
        cy, cm = _prev_month(cy, cm)
    months_back.reverse()   # eskiden yeniye

    stocks = []
    for ticker, daily in raw['daily'].items():
        if not daily: continue
        meta = raw['tickers'].get(ticker, {'type':'Stock','industry':'','name':ticker})

        mom_12_1  = calc_momentum_12_1(daily, last_date)
        fip_annual = calc_fip_annual(daily, last_date)

        # Aylık tarihçe
        monthly = []
        for (my, mm) in months_back:
            mom_m = calc_monthly_momentum(daily, my, mm)
            fip_m = calc_fip_monthly(daily, my, mm)
            monthly.append({
                'month':    f"{my}-{mm:02d}",
                'momentum': round(mom_m, 6) if mom_m is not None else None,
                'fip':      round(fip_m, 6) if fip_m is not None else None,
            })

        stocks.append({
            'ticker':   ticker,
            'name':     meta['name'],
            'type':     meta['type'],
            'industry': meta['industry'],
            'mom':      round(mom_12_1,  6) if mom_12_1  is not None else None,
            'fip':      round(fip_annual, 6) if fip_annual is not None else None,
            'monthly':  monthly,
        })

    stocks.sort(key=lambda s: s['mom'] if s['mom'] is not None else -999, reverse=True)
    return stocks

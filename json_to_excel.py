"""
json_to_excel.py — JSON dosyalarındaki güncel veriyi Excel'e yazar.

JSON'lar güncel ama Excel güncellenmemişse bu scripti çalıştır.
docs/data/detail/*.json dosyalarını okur, Excel'deki ilgili hücreleri doldurur.

Kullanım:
    python json_to_excel.py              # Önce rapor — ne yazılacak göster
    python json_to_excel.py --fix        # Excel'i güncelle
"""

import argparse
import datetime
import json
import logging
import sys
from pathlib import Path

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s  %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger(__name__)

EXCEL_FILE  = 'Claude_Momentum_Screener_BIST_Labyrinth.xlsx'
JSON_DIR    = Path('docs/data/detail')
SHEET_RAW   = 'BIST D Return Data'
ROW_DATES   = 3
ROW_START   = 6
COL_TICKER  = 4   # D sütunu


def load_json_data(json_dir: Path) -> dict:
    """
    Tüm JSON dosyalarını okur.
    Döner: {ticker: {date: ret_decimal}}
    """
    all_data = {}
    files = list(json_dir.glob('*.json'))
    log.info(f"{len(files)} JSON dosyası bulundu.")

    for f in sorted(files):
        try:
            data = json.loads(f.read_text(encoding='utf-8'))
            ticker = data.get('ticker')
            if not ticker:
                continue
            daily = {}
            for month in data.get('monthly', []):
                for day in month.get('days', []):
                    try:
                        d = datetime.date.fromisoformat(day['date'])
                        daily[d] = day['ret']
                    except:
                        pass
            if daily:
                all_data[ticker] = daily
        except Exception as e:
            log.warning(f"  {f.name}: {e}")

    log.info(f"{len(all_data)} hissenin JSON verisi yüklendi.")
    return all_data


def load_excel_structure(excel_path: Path):
    """
    Excel'deki tarih→sütun ve ticker→satır eşlemelerini okur.
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[SHEET_RAW]

    date_row    = list(ws.iter_rows(min_row=ROW_DATES, max_row=ROW_DATES, values_only=True))[0]
    date_to_col = {}
    for i, v in enumerate(date_row):
        if isinstance(v, datetime.datetime):
            date_to_col[v.date()] = i + 1  # 1-indexed

    ticker_to_row = {}
    for ri, row in enumerate(ws.iter_rows(min_row=ROW_START, max_row=700, values_only=True)):
        t = row[COL_TICKER - 1]
        if t:
            ticker_to_row[str(t).strip()] = ri + ROW_START

    wb.close()
    log.info(f"Excel yapısı: {len(date_to_col)} tarih sütunu, {len(ticker_to_row)} hisse satırı")
    return date_to_col, ticker_to_row


def find_missing(json_data: dict, date_to_col: dict, ticker_to_row: dict, excel_path: Path):
    """
    JSON'da olup Excel'de boş olan hücreleri bulur.
    Döner: {date: {ticker: ret}}
    """
    # Excel'deki mevcut değerleri oku
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[SHEET_RAW]

    # Tüm veriyi belleğe al (hızlı erişim için)
    excel_values = {}
    for row in ws.iter_rows(min_row=ROW_START, max_row=700, values_only=True):
        t = row[COL_TICKER - 1]
        if not t:
            continue
        ticker = str(t).strip()
        excel_values[ticker] = {}
        for ci, val in enumerate(row):
            excel_values[ticker][ci + 1] = val  # col 1-indexed
    wb.close()

    to_write = {}
    total_missing = 0

    for ticker, daily in json_data.items():
        if ticker not in ticker_to_row:
            continue
        for date, ret in daily.items():
            col = date_to_col.get(date)
            if col is None:
                continue  # Bu tarih Excel'de yok (Excel'in tarih aralığı dışında)
            # Mevcut Excel değeri
            current = excel_values.get(ticker, {}).get(col)
            if current is None or current == '' or current == 0:
                to_write.setdefault(date, {})[ticker] = ret
                total_missing += 1

    log.info(f"Excel'de boş olan: {total_missing} hisse-gün kaydı ({len(to_write)} farklı tarih)")
    return to_write


def write_to_excel(to_write: dict, date_to_col: dict, ticker_to_row: dict, excel_path: Path):
    """
    Eksik verileri Excel'e yazar.
    """
    if not to_write:
        log.info("Yazılacak veri yok.")
        return

    # Backup
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    backup = excel_path.with_name(f"{excel_path.stem}_backup_{ts}{excel_path.suffix}")
    try:
        backup.write_bytes(excel_path.read_bytes())
        log.info(f"Backup: {backup.name}")
    except Exception as e:
        log.warning(f"Backup atlandı: {e}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[SHEET_RAW]

    written = 0
    skipped = 0
    for date, day_data in sorted(to_write.items()):
        col = date_to_col.get(date)
        if col is None:
            skipped += len(day_data)
            continue
        for ticker, ret in day_data.items():
            row = ticker_to_row.get(ticker)
            if row is None:
                skipped += 1
                continue
            # ret decimal → yüzde (Excel formatı: -1.039 gibi)
            ws.cell(row=row, column=col).value = round(ret * 100, 8)
            written += 1

    wb.save(excel_path)
    wb.close()
    log.info(f"✅ Yazılan: {written} hücre, atlanan: {skipped}")


def main():
    parser = argparse.ArgumentParser(description='JSON → Excel senkronizasyonu')
    parser.add_argument('--fix', action='store_true', help='Excel\'i güncelle (yoksa sadece rapor)')
    args = parser.parse_args()

    excel = Path(EXCEL_FILE)
    if not excel.exists():
        log.error(f"Excel bulunamadı: {excel}")
        sys.exit(1)

    if not JSON_DIR.exists():
        log.error(f"JSON klasörü bulunamadı: {JSON_DIR}")
        sys.exit(1)

    log.info("JSON dosyaları yükleniyor...")
    json_data = load_json_data(JSON_DIR)

    log.info("Excel yapısı okunuyor...")
    date_to_col, ticker_to_row = load_excel_structure(excel)

    log.info("Eksikler hesaplanıyor...")
    to_write = find_missing(json_data, date_to_col, ticker_to_row, excel)

    if not to_write:
        log.info("✅ Excel zaten güncel.")
        return

    if args.fix:
        log.info("--fix modu: Excel güncelleniyor...")
        write_to_excel(to_write, date_to_col, ticker_to_row, excel)
        log.info("Tamamlandı. Siteyi yeniden derlemek için: python run.py")
    else:
        log.info("Rapor modu — Excel'i güncellemek için: python json_to_excel.py --fix")


if __name__ == '__main__':
    main()

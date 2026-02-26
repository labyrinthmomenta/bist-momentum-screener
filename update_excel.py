"""
update_excel.py — Excel'e yeni günlerin verisini yazar.
Sunucuda ve Mac'te çalışır.
"""
import datetime, math, logging
from pathlib import Path
import openpyxl

log = logging.getLogger(__name__)

SHEET_RAW  = "BIST D Return Data"
ROW_DATES  = 3
ROW_START  = 6
COL_TICKER = 4   # D sütunu (1-indexed)


def write_new_data(excel_path: Path, fetched: dict) -> bool:
    """
    fetched: {date: {ticker: pct_decimal}}
    Excel'e yazar. Başarı: True
    """
    if not fetched:
        log.info("Yazılacak veri yok.")
        return True

    wb  = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws  = wb[SHEET_RAW]

    # Tarih → sütun index (0-based)
    date_row    = list(ws.iter_rows(min_row=ROW_DATES, max_row=ROW_DATES, values_only=True))[0]
    date_to_col = {
        v.date(): i
        for i, v in enumerate(date_row)
        if isinstance(v, datetime.datetime)
    }

    # Ticker → satır
    ticker_to_row = {}
    for ri, row in enumerate(ws.iter_rows(min_row=ROW_START, max_row=650, values_only=True)):
        t = row[COL_TICKER - 1]
        if t:
            ticker_to_row[str(t).strip()] = ri + ROW_START
    wb.close()

    # Backup
    ts     = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = excel_path.with_name(f"{excel_path.stem}_backup_{ts}{excel_path.suffix}")
    try:
        with open(excel_path, "rb") as s, open(backup, "wb") as d:
            while chunk := s.read(1024 * 1024):
                d.write(chunk)
        log.info(f"Backup: {backup.name}")
    except Exception as e:
        log.warning(f"Backup atlandı: {e}")

    # Yaz
    wb2 = openpyxl.load_workbook(excel_path)
    ws2 = wb2[SHEET_RAW]

    written = skipped = 0
    for date, day_data in sorted(fetched.items()):
        ci = date_to_col.get(date)
        if ci is None:
            log.warning(f"{date} Excel'de yok")
            continue
        excel_col = ci + 1
        for ticker, val in day_data.items():
            row = ticker_to_row.get(ticker)
            if row is None: continue
            # pct_decimal → yüzde (Excel'deki format: -1.039 gibi)
            ws2.cell(row=row, column=excel_col).value = round(val * 100, 8)
            written += 1

    log.info(f"Yazılan hücre: {written}")
    wb2.save(excel_path)
    wb2.close()
    return True
